import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from geopy.distance import geodesic
from streamlit_js_eval import get_geolocation
import time
from datetime import datetime, timedelta
import pytz
from modules import *
from io import BytesIO
from openpyxl.styles import Alignment, Border, Side, PatternFill

LAB_LAT = 37.456461 
LAB_LON = 126.952096 
ALLOWED_RADIUS_M = 100 

st.set_page_config(page_title="출결 체크", page_icon="📍", initial_sidebar_state="collapsed")
if 'current_view' not in st.session_state:
    st.session_state['current_view'] = 'main'
if 'show_late_dialog' not in st.session_state:
    st.session_state['show_late_dialog'] = False
if 'show_early_leave_dialog' not in st.session_state:
    st.session_state['show_early_leave_dialog'] = False
if 'show_absent_dialog' not in st.session_state:
    st.session_state['show_absent_dialog'] = False

def set_view(view_name):
    st.session_state['current_view'] = view_name
    st.rerun()

# --- 다이얼로그 및 헬퍼 함수 ---
if hasattr(st, "dialog"): dlg = st.dialog
else: dlg = st.experimental_dialog

@dlg("조퇴 확인")
def show_early_leave_dialog(name, user_lat, user_lon, distance):
    st.warning("⚠️ 현재 오후 6시 이전입니다. 조퇴하시겠습니까?")
    # 조퇴 사유 입력
    reason = st.text_area(
        "조퇴 사유",
        placeholder="예: 병원 예약, 가족 행사, 개인 사정 등",
        help="조퇴 사유를 간단히 입력해주세요.",
    )
    col_y, col_n = st.columns(2)
    with col_y:
        if st.button("네 (조퇴)"):
            try:
                if not reason or not reason.strip():
                    st.warning("조퇴 사유를 입력해주세요.")
                    st.stop()
                sheet = get_sheet()
                kst = pytz.timezone('Asia/Seoul')
                now = datetime.now(kst).strftime('%Y-%m-%d %H:%M:%S')
                # 조퇴 사유는 기존대로. (스키마상 6번째 컬럼 추정)
                sheet.append_row([now, name, "조퇴", f"{user_lat},{user_lon}", f"{distance:.1f}m", reason.strip()])
                clear_attendance_cache()
                st.success(f"{name}님 {now} 조퇴 기록 완료!")
                st.balloons()
                st.session_state['show_early_leave_dialog'] = False
                time.sleep(1.5)
                st.rerun()
            except Exception as e:
                import traceback
                st.code(traceback.format_exc())
    with col_n:
        if st.button("아니오"):
            st.session_state['show_early_leave_dialog'] = False
            st.rerun()

@dlg("지각 확인")
def show_late_dialog(name, user_lat, user_lon, distance):
    st.warning("⚠️ 현재 오전 10시 이후입니다. 지각 사유를 작성해주세요.")
    # 지각 사유 입력
    reason = st.text_area(
        "지각 사유",
        placeholder="예: [업무] 외근 복귀, 병원 진료 등",
        help="지각 사유를 입력해주세요. [업무]를 포함하면 근무로 인정됩니다.",
    )
    col_y, col_n = st.columns(2)
    with col_y:
        if st.button("네 (지각 출근)"):
            try:
                if not reason or not reason.strip():
                    st.warning("지각 사유를 입력해주세요.")
                    st.stop()
                sheet = get_sheet()
                kst = pytz.timezone('Asia/Seoul')
                now = datetime.now(kst).strftime('%Y-%m-%d %H:%M:%S')
                # 스키마: 날짜, 이름, 상태, 위치, 거리, 조퇴사유, 지각사유, 결근사유
                # 지각사유는 7번째(index 6)이므로 앞의 조퇴사유(index 5)는 빈값 처리
                sheet.append_row([now, name, "지각", f"{user_lat},{user_lon}", f"{distance:.1f}m", "", reason.strip()])
                clear_attendance_cache()
                st.success(f"{name}님 {now} 지각 기록 완료!")
                st.balloons()
                st.session_state['show_late_dialog'] = False
                time.sleep(1.5)
                st.rerun()
            except Exception as e:
                import traceback
                st.code(traceback.format_exc())
    with col_n:
        if st.button("아니오"):
            st.session_state['show_late_dialog'] = False
            st.rerun()

@dlg("결근 확인")
def show_absent_dialog(name):
    st.warning("결근 사유를 작성해주세요.")
    reason = st.text_area(
        "결근 사유",
        placeholder="예: 연차, 병가, 예비군 등",
        help="결근 사유를 필수로 입력해주세요.",
    )
    col_y, col_n = st.columns(2)
    with col_y:
        if st.button("네 (결근)"):
            try:
                if not reason or not reason.strip():
                    st.warning("결근 사유를 입력해주세요.")
                    st.stop()
                sheet = get_sheet()
                kst = pytz.timezone('Asia/Seoul')
                now = datetime.now(kst).strftime('%Y-%m-%d %H:%M:%S')
                # 스키마: 날짜, 이름, 상태, 위치, 거리, 조퇴사유, 지각사유, 결근사유
                # 결근사유는 8번째(index 7)
                sheet.append_row([now, name, "결근", "", "", "", "", reason.strip()])
                clear_attendance_cache()
                st.success(f"{name}님 {now} 결근 기록 완료!")
                st.balloons()
                st.session_state['show_absent_dialog'] = False
                time.sleep(1.5)
                st.rerun()
            except Exception as e:
                import traceback
                st.code(traceback.format_exc())
    with col_n:
        if st.button("취소"):
            st.session_state['show_absent_dialog'] = False
            st.rerun()

@dlg("출결 인원 선택")
def show_name_selection_dialog(user_list):
    st.write("본인의 이름을 터치해주세요.")
    cols = st.columns(3)
    for i, user in enumerate(user_list):
        display_name = user.replace("/", "\n")
        with cols[i % 3]:
            if st.button(display_name, use_container_width=True, key=f"btn_user_select_{i}"):
                st.session_state["selected_name_radio"] = user 
                st.rerun()

def process_nn_records(sheet):
    kst = pytz.timezone('Asia/Seoul')
    now_kst = datetime.now(kst)
    today = now_kst.date()
    current_hour = now_kst.hour
    data = get_cached_records(sheet)
    if not data or len(data) < 2:
        return
    headers = data[0]
    df = pd.DataFrame(data[1:], columns=headers)
    col_ts = headers[0]
    col_name = headers[1]
    col_type = headers[2]
    df['dt'] = pd.to_datetime(df[col_ts], errors='coerce')
    df = df.dropna(subset=['dt'])
    if "user_names" in st.secrets:
        all_users = st.secrets["user_names"]
    else:
        return
    nn_records = []
    for user in all_users:
        user_df = df[df[col_name] == user].copy()
        user_df['date_obj'] = user_df['dt'].dt.date
        for check_date in pd.date_range(start=today - timedelta(days=30), end=today - timedelta(days=1)):
            date_obj = check_date.date()
            if date_obj.weekday() >= 5:
                continue
            day_records = user_df[user_df['date_obj'] == date_obj]
            if day_records.empty:
                continue
            types = day_records[col_type].unique()
            has_in = "\ucd9c\uadfc" in types or "\uc9c0\uac01" in types
            has_out = "\ud1f4\uadfc" in types or "\uc870\ud1f4" in types
            has_absent = "\uacb0\uadfc" in types
            has_in_nn = "\ucd9c\uadfcNN" in types
            has_out_nn = "\ud1f4\uadfcNN" in types
            if has_absent:
                continue
            if has_in and not has_out and not has_out_nn:
                ts = datetime.combine(date_obj, datetime.min.time().replace(hour=23, minute=59))
                nn_records.append([ts.strftime('%Y-%m-%d %H:%M:%S'), user, "\ud1f4\uadfcNN", "", "", "\uc0ac\uc720\uc5c6\uc74c", "", ""])
            elif not has_in and has_out and not has_in_nn:
                ts = datetime.combine(date_obj, datetime.min.time().replace(hour=18, minute=0))
                nn_records.append([ts.strftime('%Y-%m-%d %H:%M:%S'), user, "\ucd9c\uadfcNN", "", "", "", "\uc0ac\uc720\uc5c6\uc74c", ""])
            elif not has_in and not has_out and not has_in_nn and not has_out_nn:
                ts_in = datetime.combine(date_obj, datetime.min.time().replace(hour=18, minute=0))
                ts_out = datetime.combine(date_obj, datetime.min.time().replace(hour=23, minute=59))
                nn_records.append([ts_in.strftime('%Y-%m-%d %H:%M:%S'), user, "\ucd9c\uadfcNN", "", "", "", "\uc0ac\uc720\uc5c6\uc74c", ""])
                nn_records.append([ts_out.strftime('%Y-%m-%d %H:%M:%S'), user, "\ud1f4\uadfcNN", "", "", "\uc0ac\uc720\uc5c6\uc74c", "", ""])
        if today.weekday() < 5:
            day_records = user_df[user_df['date_obj'] == today]
            if not day_records.empty:
                types = day_records[col_type].unique()
                has_in = "\ucd9c\uadfc" in types or "\uc9c0\uac01" in types
                has_out = "\ud1f4\uadfc" in types or "\uc870\ud1f4" in types
                has_absent = "\uacb0\uadfc" in types
                has_in_nn = "\ucd9c\uadfcNN" in types
                if not has_absent and not has_in and has_out and current_hour >= 18 and not has_in_nn:
                    ts = datetime.combine(today, datetime.min.time().replace(hour=18, minute=0))
                    nn_records.append([ts.strftime('%Y-%m-%d %H:%M:%S'), user, "\ucd9c\uadfcNN", "", "", "", "\uc0ac\uc720\uc5c6\uc74c", ""])
    if nn_records:
        for record in nn_records:
            sheet.append_row(record)
        clear_attendance_cache()

# --- 출결 기록 확인 페이지 ---
def view_records_page():
    st.markdown("""
    <style>
    div[data-testid="stButton"] button {
        width: 100%;
    }
    th, td {
        white-space: pre-wrap !important; 
        vertical-align: top !important;
        font-size: 0.9rem !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.title("📋 출결 현황")

    try:
        sheet = get_sheet()
        process_nn_records(sheet)
        data = get_cached_records(sheet)
        
        if data and len(data) > 1:
            headers = data[0]
            df = pd.DataFrame(data[1:], columns=headers)
            
            # 컬럼 인덱스 매핑 (헤더 이름이 바뀔 수도 있으므로 위치 기반 추정 혹은 이름 확인)
            # 0:timestamp, 1:name, 2:type, 3:loc, 4:dist, 5:early_reason, 6:late_reason, 7:absent_reason
            col_ts = headers[0] 
            col_name = headers[1] 
            col_type = headers[2] 
            
            df['dt'] = pd.to_datetime(df[col_ts], errors='coerce')
            df = df.dropna(subset=['dt'])
            
            kst = pytz.timezone('Asia/Seoul')
            now_kst = datetime.now(kst)
            today = now_kst.date()
            
            mask_month = (df['dt'].dt.year == today.year) & (df['dt'].dt.month == today.month)
            month_df = df[mask_month].copy().sort_values('dt')
            
            all_users = sorted(month_df[col_name].unique())

            # --- Container 1: 개인별 현황 리스트 ---
            with st.container(border=True):
                st.subheader("👤 개인별 현황 상세")
                
                # 기본 선택값 설정
                default_idx = 0
                if st.session_state.get("selected_name_radio") in all_users:
                    default_idx = all_users.index(st.session_state["selected_name_radio"])
                selected_user = st.selectbox("이름을 선택하세요", all_users, index=default_idx)
                
                if selected_user:
                    user_df = month_df[month_df[col_name] == selected_user].copy()
                    # 보여줄 데이터 가공
                    display_list = []
                    for _, row in user_df.iterrows():
                        r_type = row[col_type]
                        r_dt = row['dt']
                        
                        reason = ""
                        vals = row.values
                        if r_type == "조퇴" and len(vals) > 5:
                            reason = str(vals[5])
                        elif r_type == "지각" and len(vals) > 6:
                            reason = str(vals[6])
                        elif r_type == "결근" and len(vals) > 7:
                            reason = str(vals[7])
                            
                        display_list.append({
                            "날짜": r_dt.strftime("%Y-%m-%d"),
                            "시간": r_dt.strftime("%H:%M:%S"),
                            "구분": r_type,
                            "내용/사유": reason
                        })
                    
                    st.dataframe(pd.DataFrame(display_list), use_container_width=True, hide_index=True)

            # --- Container 2: 월간 현황 다운로드 ---
            with st.container(border=True):
                col_h1, col_h2 = st.columns([3, 1])
                with col_h1:
                    st.subheader("📥 월간 전체 현황 다운로드")
                with col_h2:
                    if st.button("🔄 새로고침"):
                        clear_attendance_cache()
                        st.rerun()
                import calendar
                
                _, last_day = calendar.monthrange(today.year, today.month)
                valid_date_cols = []
                valid_dates = []
                
                for d in range(1, last_day + 1):
                    curr_date = datetime(today.year, today.month, d).date()
                    if curr_date.weekday() < 5: # 0(월)~4(금)
                        valid_dates.append(curr_date)
                        valid_date_cols.append(f"{d}일")

                report_data = []

                for u in all_users:
                    u_df = month_df[month_df[col_name] == u]
                    late_cnt = early_cnt = absent_cnt = 0
                    day_status_map = {}
                    u_df['date_obj'] = u_df['dt'].dt.date
                    grouped = u_df.groupby('date_obj')
                    for d_date, grp in grouped:
                        types = grp[col_type].unique()
                        parts = []
                        in_time = out_time = ""
                        notes = []
                        has_work = False
                        has_in_nn = has_out_nn = False
                        if "출근" in types:
                            in_time = grp[grp[col_type]=="출근"]['dt'].min().strftime("%H:%M")
                            has_work = True
                        elif "출근NN" in types:
                            in_time = "NN"
                            has_in_nn = True
                            has_work = True
                        if "지각" in types:
                            late_rows = grp[grp[col_type]=="지각"]
                            is_excused = False
                            reason_txt = ""
                            for _, r in late_rows.iterrows():
                                if len(r) > 6:
                                    r_val = str(r.iloc[6]) if pd.notna(r.iloc[6]) else ""
                                    reason_txt = r_val
                                    if "[업무]" in r_val:
                                        is_excused = True
                            if not reason_txt or reason_txt.lower() in ["nan","none",""]:
                                reason_txt = "사유없음"
                            if not is_excused:
                                late_cnt += 1
                            in_time = grp[grp[col_type]=="지각"]['dt'].min().strftime("%H:%M")
                            notes.append(f"지각({reason_txt})" if not is_excused else f"지각(업무:{reason_txt})")
                            has_work = True
                        if "퇴근" in types:
                            out_time = grp[grp[col_type]=="퇴근"]['dt'].max().strftime("%H:%M")
                        elif "퇴근NN" in types:
                            out_time = "NN"
                            has_out_nn = True
                        if "조퇴" in types:
                            early_rows = grp[grp[col_type]=="조퇴"]
                            is_excused = False
                            reason_txt = ""
                            for _, r in early_rows.iterrows():
                                if len(r) > 5:
                                    r_val = str(r.iloc[5]) if pd.notna(r.iloc[5]) else ""
                                    reason_txt = r_val
                                    if "[업무]" in r_val:
                                        is_excused = True
                            if not reason_txt or reason_txt.lower() in ["nan","none",""]:
                                reason_txt = "사유없음"
                            if not is_excused:
                                early_cnt += 1
                            out_time = grp[grp[col_type]=="조퇴"]['dt'].max().strftime("%H:%M")
                            notes.append(f"조퇴({reason_txt})" if not is_excused else f"조퇴(업무:{reason_txt})")
                        if "결근" in types:
                            absent_cnt += 1
                            abs_rows = grp[grp[col_type]=="결근"]
                            reason_txt = ""
                            if not abs_rows.empty:
                                r = abs_rows.iloc[0]
                                if len(r) > 7:
                                    reason_txt = str(r.iloc[7]) if pd.notna(r.iloc[7]) else ""
                            if not reason_txt or reason_txt.lower() in ["nan","none",""]:
                                reason_txt = "사유없음"
                            notes.append(f"결근({reason_txt})")
                        if has_in_nn:
                            late_cnt += 1
                            nn_reason = ""
                            nn_rows = grp[grp[col_type]=="출근NN"]
                            if not nn_rows.empty:
                                r = nn_rows.iloc[0]
                                if len(r) > 6:
                                    nn_reason = str(r.iloc[6]) if pd.notna(r.iloc[6]) else "사유없음"
                            if not nn_reason or nn_reason.lower() in ["nan","none",""]:
                                nn_reason = "사유없음"
                            notes.insert(0, f"지각({nn_reason})")
                        if has_out_nn:
                            early_cnt += 1
                            nn_reason = ""
                            nn_rows = grp[grp[col_type]=="퇴근NN"]
                            if not nn_rows.empty:
                                r = nn_rows.iloc[0]
                                if len(r) > 5:
                                    nn_reason = str(r.iloc[5]) if pd.notna(r.iloc[5]) else "사유없음"
                            if not nn_reason or nn_reason.lower() in ["nan","none",""]:
                                nn_reason = "사유없음"
                            notes.append(f"조퇴({nn_reason})")
                        if has_in_nn and has_out_nn:
                            late_cnt -= 1
                            early_cnt -= 1
                            absent_cnt += 1
                            notes = ["결근(사유없음)"]
                        if in_time:
                            parts.append(f"출근: {in_time}")
                        if out_time:
                            parts.append(f"퇴근: {out_time}")
                        elif has_work and "결근" not in types and not has_out_nn:
                            parts.append("퇴근: NN")
                        if notes:
                            parts.append(", ".join(notes))
                        day_status_map[d_date] = ", ".join(parts)
                    row = {
                        "이름": u,
                        "현황": f"결근:{absent_cnt}, 지각:{late_cnt}, 조퇴:{early_cnt}"
                    }
                    for d_date, d_col in zip(valid_dates, valid_date_cols):
                        row[d_col] = day_status_map.get(d_date, "")
                    
                    report_data.append(row)
                
                if report_data:
                    rep_df = pd.DataFrame(report_data)
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        rep_df.to_excel(writer, index=False, sheet_name='출결현황')
                        worksheet = writer.sheets['출결현황']
                        worksheet.column_dimensions['A'].width = 20
                        worksheet.column_dimensions['B'].width = 25
                        for col in range(3, len(rep_df.columns) + 1):
                            worksheet.column_dimensions[worksheet.cell(1, col).column_letter].width = 35
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                        fill_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                        for idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=len(rep_df)+1), start=1):
                            fill = fill_white if idx % 2 == 0 else fill_gray
                            for cell in row:
                                cell.border = thin_border
                                if idx > 1:
                                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                                    cell.fill = fill
                    buffer.seek(0)
                    st.download_button(
                        label="💾 Excel 파일 다운로드",
                        data=buffer,
                        file_name=f"출결현황_{today.year}_{today.month}월.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.info("데이터가 없습니다.")

        else:
            st.info("데이터가 없습니다.")
    except Exception as e:
        st.error(f"오류 발생: {e}")
        # import traceback
        # st.code(traceback.format_exc())

    st.divider()
    if st.button("🏠 메인 화면으로 이동"):
        set_view('main')

# --- 메인 출결체크 페이지 ---
def view_main_page():
    st.markdown("""
        <style>
        .responsive-title {
            font-size: clamp(1.2rem, 5vw, 2rem);
            font-weight: bold;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            margin-bottom: 20px;
        }
        div[role="radiogroup"] > label {
            margin-bottom: 12px !important;
            padding: 10px !important;
            border-radius: 8px;
            background-color: #f0f2f6;
        }
        div[role="radiogroup"] > label:hover {
            background-color: #e0e2e6;
        }
        div[data-testid="stButton"] button p {
            white-space: pre-wrap !important;
            line-height: 1.2 !important;
            text-align: center !important;
        }
        </style>
        <div class="responsive-title">📍SCP-LAB 위치 기반 출퇴근 기록</div>
        """, unsafe_allow_html=True)

    # 페이지 이동 버튼
    if st.button("📋 전체 기록 보기", use_container_width=True):
        set_view('records')

    # 사용자 정보 확인
    if "user_names" in st.secrets:
        user_list = st.secrets["user_names"]
    else:
        user_list = ["관리자에게 문의하세요"]
    
    if "selected_name_radio" not in st.session_state:
        st.session_state["selected_name_radio"] = None
    name = st.session_state["selected_name_radio"]

    if not name:
        st.info("본인을 선택해주세요 🔽")
        if st.button("사용자 선택", use_container_width=True, type="primary"):
            show_name_selection_dialog(user_list)
    else:
        c1, c2 = st.columns([1, 5])
        with c1:
            if st.button("🔄", help="사용자 변경"):
                show_name_selection_dialog(user_list)
        with c2:
            st.success(f"**{name}**님 안녕하세요! 👋")
        
        # 결근 버튼 (위치 무관)
        is_absent = check_is_absent_today(get_sheet(), name)
        if is_absent:
            st.button("🙅 결근 통보 (위치 무관)", disabled=True, use_container_width=True)
            st.info("이미 오늘 결근 기록이 있습니다.")
        else:
            if st.button("🙅 결근 통보 (위치 무관)", use_container_width=True):
                st.session_state['show_absent_dialog'] = True
                st.rerun()
            
            if st.session_state.get('show_absent_dialog'):
                show_absent_dialog(name)
                st.session_state['show_absent_dialog'] = False

    # 위치 확인 및 출결 로직
    loc = get_geolocation()
    if loc and 'coords' in loc:
        user_lat = loc['coords']['latitude']
        user_lon = loc['coords']['longitude']
        office_point = (LAB_LAT, LAB_LON)
        user_point = (user_lat, user_lon)
        distance = geodesic(office_point, user_point).meters
        
        st.write(f"현재 위치 감지됨: 연구실과의 거리 **{distance:.1f}m**")
        
        if distance <= ALLOWED_RADIUS_M:
            st.success("✅ 연구실 근처입니다. 출퇴근이 가능합니다.")
            col1, col2 = st.columns(2)
            
            is_in = check_is_clocked_in(get_sheet(), name)
            is_out = check_is_clocked_out(get_sheet(), name)

            with col1:
                if is_in:
                    st.button("출근하기 ☀️", disabled=True, key="btn_in_disabled")
                    st.info("이미 오늘 출근 기록이 있습니다.")
                elif is_out:
                    st.button("출근하기 ☀️", disabled=True, key="btn_in_disabled_out")
                    st.info("이미 오늘 퇴근 기록이 있습니다.")
                else:
                    if st.button("출근하기 ☀️", key="btn_in_active"):
                        if not name:
                            st.warning("이름을 입력해주세요.")
                        else:
                            kst = pytz.timezone('Asia/Seoul')
                            now_dt = datetime.now(kst)
                            if now_dt.hour >= 10:
                                # 10시 이후 -> 지각 다이얼로그
                                st.session_state['show_late_dialog'] = True
                                st.rerun()
                            else:
                                # 10시 이전 -> 정상 출근 바로 기록
                                try:
                                    sheet = get_sheet()
                                    now = now_dt.strftime('%Y-%m-%d %H:%M:%S')
                                    sheet.append_row([now, name, "출근", f"{user_lat},{user_lon}", f"{distance:.1f}m", "", "", ""])
                                    clear_attendance_cache()
                                    st.success(f"{name}님 {now} 출근 기록 완료!")
                                    st.balloons()
                                    time.sleep(1.5)
                                    st.rerun()
                                except Exception as e:
                                    import traceback
                                    st.error(f"출근 기록 오류: {e}")
                                    st.code(traceback.format_exc())
                    
                    if st.session_state.get('show_late_dialog'):
                        show_late_dialog(name, user_lat, user_lon, distance)
                        st.session_state['show_late_dialog'] = False

            with col2:
                if is_out:
                    st.button("퇴근하기 🌙", disabled=True, key="btn_out_disabled")
                    st.info("이미 오늘 퇴근 하셨습니다!")
                else:
                    if st.button("퇴근하기 🌙"):
                        if not name:
                            st.warning("이름을 입력해주세요.")
                        else:
                            kst = pytz.timezone('Asia/Seoul')
                            now_dt = datetime.now(kst)
                            if now_dt.hour < 18:
                                # 18시 이전 -> 조퇴 다이얼로그
                                st.session_state['show_early_leave_dialog'] = True
                                st.rerun()
                            else:
                                # 18시 이후 -> 정상 퇴근 바로 기록
                                try:
                                    sheet = get_sheet()
                                    now = now_dt.strftime('%Y-%m-%d %H:%M:%S')
                                    sheet.append_row([now, name, "퇴근", f"{user_lat},{user_lon}", f"{distance:.1f}m", "", "", ""])
                                    clear_attendance_cache()
                                    st.success(f"{name}님 {now} 퇴근 기록 완료!")
                                    st.balloons()
                                    time.sleep(1.5)
                                    st.rerun()
                                except Exception as e:
                                    import traceback
                                    st.error(f"퇴근 기록 오류: {e}")
                                    st.code(traceback.format_exc())
                    
                    if st.session_state.get('show_early_leave_dialog'):
                        show_early_leave_dialog(name, user_lat, user_lon, distance)
                        st.session_state['show_early_leave_dialog'] = False
        else:
            st.error(f"🚫 연구실 반경 {ALLOWED_RADIUS_M}m 밖입니다. 출퇴근을 기록할 수 없습니다.")
        
        df_map = pd.DataFrame({'lat': [user_lat, LAB_LAT], 'lon': [user_lon, LAB_LON]})
        st.map(df_map, zoom=15)
    elif loc and 'error' in loc:
        st.error(f"⚠️ 위치 정보를 불러오지 못했습니다: {loc['error']}\n브라우저 '위치 권한'을 허용했는지 확인해주세요.")
    else:
        st.info("📍 위치 권한을 허용하고 잠시 기다려주세요 (브라우저 새로고침 필요할 수 있음)")

# --- 라우팅 로직 ---
if st.session_state['current_view'] == 'records':
    view_records_page()
else:
    view_main_page()

